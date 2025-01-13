import requests
import json
import pandas as pd
from openpyxl import load_workbook
import hashlib
from datetime import datetime
import random

# 엑셀 파일 경로 및 시트 이름 설정
excel_file_path = "coupontest_use.xlsx"
variable_sheet = "Variable"
results_sheet = "results"


def use_coupon_api(wb, file_path, results_sheet):
    """
    Use_coupon 함수를 호출하고 결과를 results 시트에 기록하는 함수.
    """
    # 시트 불러오기
    res_ws = wb[results_sheet]
    headers = [cell.value for cell in res_ws[1]]

    # 헤더 인덱스 설정
    coupon_history_id_idx = headers.index("Issue couponHistoryId") + 1
    coupon_id_idx = headers.index("Coupon ID") + 1
    coupon_number_idx = headers.index("Issue couponNumber") + 1

    # 추가 헤더 설정
    if "Use Status" not in headers:
        res_ws.cell(row=1, column=len(headers) + 1, value="Use Status")
        headers.append("Use Status")
    if "Use Request" not in headers:
        res_ws.cell(row=1, column=len(headers) + 1, value="Use Request")
        headers.append("Use Request")
    if "Use Response" not in headers:
        res_ws.cell(row=1, column=len(headers) + 1, value="Use Response")
        headers.append("Use Response")

    use_status_idx = headers.index("Use Status") + 1
    use_request_idx = headers.index("Use Request") + 1
    use_response_idx = headers.index("Use Response") + 1

    # 고정값
    user_idx = "373109"
    base_amount = 1000000
    coupon_goods = "1760"
    registration_by = "jun_auto"
    order_counter = 1  # order_number 증가값

    # 랜덤값 리스트
    coupon_use_member_grades = ["PRE_VIP", "VIP", "VVIP", "EXCLUSIVE_VVIP", "NONE"]
    coupon_use_sites = ["MEMBERSHIP", "CS", "HOMEPAGE"]
    coupon_goods_categories = ["MASSAGE", "LACLOUD", "W-WATER"]

    # 각 행 반복
    for row_idx in range(2, res_ws.max_row + 1):
        coupon_history_id = res_ws.cell(row=row_idx, column=coupon_history_id_idx).value
        coupon_id = res_ws.cell(row=row_idx, column=coupon_id_idx).value
        coupon_number = res_ws.cell(row=row_idx, column=coupon_number_idx).value

        if not coupon_history_id or not coupon_id or not coupon_number:
            print(f"Skipping Row {row_idx} due to missing data.")
            continue

        # order_number 생성 (고유값)
        order_number = f"{registration_by}_{order_counter}"
        order_counter += 1

        # 랜덤값 설정
        coupon_use_member_grade = random.choice(coupon_use_member_grades)
        coupon_use_site = random.choice(coupon_use_sites)
        coupon_goods_category = random.choice(coupon_goods_categories)

        # 쿠폰 사용 일시
        coupon_use_date = int(datetime.now().timestamp())

        # Hash 생성
        hash_input = f"{coupon_id}{coupon_history_id}{user_idx}{order_number}{coupon_number}{coupon_use_date}{base_amount}"
        hash_value = hashlib.sha256(hash_input.encode()).hexdigest()

        # Payload 생성
        payload = {
            "couponUseForms": [
                {
                    "couponHistoryId": coupon_history_id,
                    "couponId": coupon_id,
                    "userIdx": user_idx,
                    "couponUseMemberGrade": coupon_use_member_grade,
                    "orderNumber": order_number,
                    "couponNumber": coupon_number,
                    "baseAmount": base_amount,
                    "couponUseSite": coupon_use_site,
                    "couponGoodsCategory": coupon_goods_category,
                    "couponGoods": coupon_goods,
                    "couponUseDate": coupon_use_date,
                    "hash": hash_value,
                    "registrationBy": registration_by
                }
            ]
        }

        try:
            # API 호출
            url = "https://tcoupon.bodyfriend.com/api/coupon/history/v1/use"
            headers = {
                "accept": "application/json;charset=UTF-8",
                "code": "COUPON",
                "key": "bodyfsdf#$asdfBdf3"
            }
            response = requests.post(url, headers=headers, json=payload)
            response_json = response.json()

            # 결과 저장
            res_ws.cell(row=row_idx, column=use_status_idx, value=response.status_code)
            res_ws.cell(row=row_idx, column=use_request_idx, value=json.dumps(payload, ensure_ascii=False, indent=2))
            res_ws.cell(row=row_idx, column=use_response_idx, value=json.dumps(response_json, ensure_ascii=False, indent=2))

            print(f"Row {row_idx} processed successfully. Status: {response.status_code}")

        except Exception as e:
            print(f"Error at Row {row_idx}: {e}")
            res_ws.cell(row=row_idx, column=use_status_idx, value="ERROR")
            res_ws.cell(row=row_idx, column=use_request_idx, value=json.dumps(payload, ensure_ascii=False, indent=2))
            res_ws.cell(row=row_idx, column=use_response_idx, value=str(e))

        # 실시간 저장
        wb.save(file_path)

    print("Coupon usage process completed.")

# 실행 코드
try:
    wb = load_workbook(excel_file_path)
    use_coupon_api(wb, excel_file_path, results_sheet)
except KeyboardInterrupt:
    print("Process interrupted. Saving workbook...")
finally:
    wb.save(excel_file_path)
    wb.close()
    print("Workbook saved and closed.")
