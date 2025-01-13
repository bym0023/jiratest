import requests
import json
import pandas as pd
from openpyxl import load_workbook
import time
from itertools import product
import hashlib
from datetime import datetime
import random

# 엑셀 파일 경로 및 시트 이름 설정
excel_file_path = "coupontest_issue.xlsx"
variable_sheet = "Variable"
results_sheet = "results"
variable_list_sheet = "Variable List"  # 데이터 조합 생성용 시트 이름

def issue_coupon_from_excel(excel_file_path, results_sheet):
    """
    Reads coupon IDs from the 'results' sheet in Excel and issues coupons via API.
    Fixed and random values are used for other required parameters.

    Args:
        excel_file_path (str): Path to the Excel file.
        results_sheet (str): Name of the sheet containing coupon IDs.
    """
    # Load workbook and sheet
    wb = load_workbook(excel_file_path)
    if results_sheet not in wb.sheetnames:
        raise ValueError(f"'{results_sheet}' sheet not found in the workbook.")
    res_ws = wb[results_sheet]

    # Find or add necessary headers
    headers = [cell.value for cell in res_ws[1] if cell.value]
    if "Coupon ID" not in headers:
        raise ValueError("'Coupon ID' column not found in the results sheet.")
    
    # Ensure new headers exist
    required_headers = ["Issue Status", "Issue Request", "Issue Response", "Issue couponHistoryId", "Issue couponNumber"]
    for header in required_headers:
        if header not in headers:
            headers.append(header)
            res_ws.cell(row=1, column=len(headers), value=header)
    
    # Update header indexes
    coupon_id_col_idx = headers.index("Coupon ID") + 1
    issue_status_col_idx = headers.index("Issue Status") + 1
    issue_request_col_idx = headers.index("Issue Request") + 1
    issue_response_col_idx = headers.index("Issue Response") + 1
    issue_coupon_history_id_col_idx = headers.index("Issue couponHistoryId") + 1
    issue_coupon_number_col_idx = headers.index("Issue couponNumber") + 1


    # Fixed and random values
    user_idx = "373109"  # Fixed value
    issue_date = 1734400671  # Fixed value (timestamp)
    registration_by = "jun_auto"  # Fixed value
    member_grades = ["PRE_VIP", "VIP", "VVIP", "EXCLUSIVE_VVIP", "NONE"]  # Random selection
    issue_sites = ["MEMBERSHIP", "CS", "HOMEPAGE"]  # Random selection

    # API URL and Headers
    url = "https://tcoupon.bodyfriend.com/api/coupon/history/v1/issue"
    headers_api = {
        "accept": "application/json;charset=UTF-8",
        "code": "COUPON",
        "key": "bodyfsdf#$asdfBdf3"
    }

    # Iterate over rows in the 'results' sheet
    last_row = res_ws.max_row
    for row_idx in range(2, last_row + 1):  # Start from the second row (skip headers)
        coupon_id = res_ws.cell(row=row_idx, column=coupon_id_col_idx).value

        if not coupon_id:  # Skip if coupon_id is empty
            continue

        # Generate random values for member grade and issue site
        member_grade = random.choice(member_grades)
        issue_site = random.choice(issue_sites)

        # Generate hash value: sha256(couponId + userIdx + couponIssueDate)
        hash_input = f"{coupon_id}{user_idx}{issue_date}"
        hash_value = hashlib.sha256(hash_input.encode()).hexdigest()

        # Prepare payload
        payload = {
            "couponId": coupon_id,
            "userIdx": user_idx,
            "couponIssueDate": issue_date,
            "registrationBy": registration_by,
            "couponIssueMemberGrade": member_grade,
            "couponIssueSite": issue_site,
            "hash": hash_value
        }

        try:
            # API Request
            response = requests.post(url, headers=headers_api, json=payload)
            response_json = response.json()
            status_code = response.status_code

            # Log the response
            print(f"[{datetime.now()}] Row {row_idx - 1}: Coupon Issued | Status: {status_code}")
            # response print
            # print(json.dumps(response_json, indent=2, ensure_ascii=False))

            # 응답 결과에서 couponHistoryId와 couponNumber 추출
            coupon_history_id = response_json.get("data", {}).get("couponHistoryId", "N/A")
            coupon_number = response_json.get("data", {}).get("couponNumber", "N/A")


            # Write results to Excel
            res_ws.cell(row=row_idx, column=issue_status_col_idx, value=status_code)
            res_ws.cell(row=row_idx, column=issue_request_col_idx, value=json.dumps(payload, indent=2, ensure_ascii=False))
            res_ws.cell(row=row_idx, column=issue_response_col_idx, value=json.dumps(response_json, indent=2, ensure_ascii=False))
            res_ws.cell(row=row_idx, column=issue_coupon_history_id_col_idx, value=coupon_history_id)
            res_ws.cell(row=row_idx, column=issue_coupon_number_col_idx, value=coupon_number)


        except Exception as e:
            print(f"Error at Row {row_idx - 1}: {e}")
            res_ws.cell(row=row_idx, column=issue_status_col_idx, value="ERROR")
            res_ws.cell(row=row_idx, column=issue_request_col_idx, value=json.dumps(payload, ensure_ascii=False))
            res_ws.cell(row=row_idx, column=issue_response_col_idx, value=str(e))
            res_ws.cell(row=row_idx, column=issue_coupon_history_id_col_idx, value="N/A")
            res_ws.cell(row=row_idx, column=issue_coupon_number_col_idx, value="N/A")

        # Save progress after each row
        wb.save(excel_file_path)

    print("Coupon issuance process completed.")


issue_coupon_from_excel(excel_file_path, results_sheet)
