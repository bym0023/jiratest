import requests
import json
import pandas as pd
from openpyxl import load_workbook
import time
from itertools import product
import hashlib
from datetime import datetime


# 엑셀 파일 경로 및 시트 이름 설정
excel_file_path = "coupontest_variable.xlsx"
variable_sheet = "Variable"
results_sheet = "results"
variable_list_sheet = "Variable List"  # 데이터 조합 생성용 시트 이름

def clear_sheet_content(sheet):
    """
    Clear all content from the given openpyxl worksheet except the header row.
    """
    # 모든 행을 삭제하기 전에 헤더 값 저장
    header = [cell.value for cell in sheet[1]]  # 첫 번째 행만 저장
    
    # 모든 행 삭제
    sheet.delete_rows(2, sheet.max_row)  # 2행부터 마지막 행까지 삭제
    
    # 헤더 복구 (헤더가 없어질 경우 대비)
    for col_num, value in enumerate(header, 1):  # 열 번호는 1부터 시작
        sheet.cell(row=1, column=col_num, value=value)

def generate_combinations(wb, variable_list_sheet, variable_sheet):
    # Variable List 시트 읽기
    if variable_list_sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{variable_list_sheet}' not found in the workbook.")
    variable_list_df = pd.read_excel(excel_file_path, sheet_name=variable_list_sheet)

    # Variable 시트 읽기
    if variable_sheet not in wb.sheetnames:
        ws = wb.create_sheet(variable_sheet)
        variable_headers = list(variable_list_df.columns)
        variable_headers.extend(["couponGoodsList", "couponGoodsNoneList"])
        ws.append(variable_headers)
    else:
        ws = wb[variable_sheet]
        variable_headers = [cell.value for cell in ws[1]]
        if not variable_headers:
            raise ValueError(f"'Variable' sheet is missing headers.")
        clear_sheet_content(ws)

    # Variable List 시트의 헤더와 Variable 시트 헤더의 공통 부분만 사용
    common_headers = [header for header in variable_list_df.columns if header in variable_headers]
    if not common_headers:
        raise ValueError("No matching headers found between 'Variable' and 'Variable List' sheets.")

    column_values = {header: variable_list_df[header].dropna().tolist() for header in common_headers}

    # 모든 조합 생성
    if column_values:
        combinations = list(product(*column_values.values()))
    else:
        combinations = []

    # 디버깅 정보 출력
    print(f"Common Headers: {list(column_values.keys())}")  # 공통 헤더 출력
    print(f"Number of Combinations: {len(combinations)}")

    start_row = ws.max_row + 1

    for combination in combinations:
        row_data = {header: value for header, value in zip(common_headers, combination)}
        # 동적 변수 생성 및 row_data 입력

        #2. 쿠폰 타입 couponType 처리 (LIST필수) - 쿠폰이름때문에 앞에 설정
        # 디스카운트, 언, 리로케이션 등
        couponType = row_data.get("couponType", "")

        #3.쿠폰 적용 방식 couponApplyMode 처리 (LIST필수) - 쿠폰이름때문에 앞에 설정
        couponApplyMode = row_data.get("couponApplyMode", "")

        #31. 요청자 registrationBy 처리 (필수) - 쿠폰이름때문에 앞에 설정
        registrationBy = "jun"
        row_data["registrationBy"] = registrationBy

        #1. 이름 couponName 처리 (필수)
        couponName = f"{registrationBy}{couponType}{couponApplyMode}"
        row_data["couponName"] = couponName

        #쿠폰 설명
        #쿠폰 표시 텍스트

        #쿠폰번호 타입 (쿠폰발급타입=NUMBER면 필수)
        #쿠폰번호 (쿠폰발급타입=NUMBER면 필수)
        #쿠폰 번호 발행수량(쿠폰발급타입=PLURLA면 필수)


        #4. 적용 금액 couponApplyAmount 처리 (필수)
        couponApplyAmount = 10 if couponApplyMode == "PERCENT" else (10000 if couponApplyMode == "AMOUNT" else "")
        row_data["couponApplyAmount"] = couponApplyAmount

        #5. 중복 발급 가능 여부 처리 (필수) - fixed_values로 관리
        #6. 중복 사용 가능 여부 처리 (필수) - fixed_values로 관리

        #7. 사용 가능 제품 카테고리 TRUE/FALSE로 변환 후 처리 (LIST필수)
        if "couponGoodsCategoryStatus" in row_data:
            row_data["couponGoodsCategoryStatus"] = (True if row_data["couponGoodsCategoryStatus"] in [True, "TRUE", 1] else False)

        #8. 제품 카테고리 목록 (카테고리 TRUE 필수)  - fixed_values로 관리

        #9. 사용 가능 제품 설정 TRUE, FALSE 처리 (LIST필수)
        row_data["couponGoodsStatus"] = (True if row_data.get("couponGoodsStatus") in [True, "TRUE", 1, "1"] else False)

        #10. 제품 목록  처리 (가능 제품 설정 TRUE면 필수)  - fixed_values로 관리

        #11. 사용 가능 제외 제품 설정 TRUE, FALSE 처리 (LIST필수)
        # row_data["couponGoodsNoneStatus"] = "TRUE" if row_data.get("couponGoodsNoneStatus") in [True, "TRUE", 1, "1"] else "FALSE"
        row_data["couponGoodsNoneStatus"] = (True if row_data.get("couponGoodsNoneStatus") in [True, "TRUE", 1, "1"] else False)

        #12. 사용 가능 제외 제품 목록 couponGoodsNoneList 처리 (제외 설정 TRUE필수)   - fixed_values로 관리

        #13. 발급 타입 설정 (필수) - LIST필수값이나 일단 다운로드만 해보는걸로.
        couponIssueType = row_data.get("couponIssueType", "")

        #발급대상 (쿠폰타입SYSTEM이면 필수)
        #발급시점타입 (쿠폰타입SYSTEM이면 필수)
        #발급시점 (쿠폰타입SYSTEM이면 필수)

        #14. 발급 가능 여부 (LIST필수)
        row_data["couponIssueStatus"] = (True if row_data.get("couponIssueStatus") in [True, "TRUE", 1, "1"] else False)

        #15. 발급 가능처 설정 여부 (LIST필수)
        row_data["couponIssueSiteStatus"] = (True if row_data.get("couponIssueSiteStatus") in [True, "TRUE", 1, "1"] else False)

        #16. 발급가능처 목록 (가능처 설정 TRUE면 필수) - fixed_values로 관리

        #쿠폰 발급 고객별 횟수 제한 couponIssueCustomerLimit

        # 16-2 발급 전체 횟수 제한 처리 (선택) - fixed_values로 관리

        #17. 고객 등급 설정 여부 (LIST필수)
        row_data["couponMemberGradeStatus"] = (True if row_data.get("couponMemberGradeStatus") in [True, "TRUE", 1, "1"] else False)

        #18. 고객 등급 목록 (등급 설정 TRUE면 필수) - fixed_values로 관리

        #19. 발급 가능 일자 설정 타입 (LIST필수)
        # couponIssueSetType = row_data.get("couponIssueSetType", "")
        couponIssueSetType = row_data.get("couponIssueSetType", "")
        row_data["couponIssueSetType"] = couponIssueSetType

        #20. 발급 시작/종료 일시 (발급 가능일자 DATE면 필수)
        if couponIssueSetType == "DATE":
            row_data["couponIssueStartAt"] = 1733400000
            row_data["couponIssueEndAt"] = 1735646400
        else:
            row_data["couponIssueStartAt"] = None
            row_data["couponIssueEndAt"] = None        


        #21. 사용 가능 여부 설정 (LIST필수)
        row_data["couponUseStatus"] = (True if row_data.get("couponUseStatus") in [True, "TRUE", 1, "1"] else False)

        #22. 사용 가능일자 설정 타입 (필수) 
        couponUseSetType = row_data.get("couponUseSetType", "")
        row_data["couponUseSetType"] = couponUseSetType


        # 사용 '시작/종료' 일시 및 사용 기간 설정
        if couponUseSetType == "DATE":
            row_data["couponUseStartAt"] = 1733400000
            row_data["couponUseEndAt"] = 1735646400
            row_data["couponUseTerm"] = None  # DATE일 경우 사용 기간은 NULL
        elif couponUseSetType == "TERM":
            row_data["couponUseStartAt"] = None
            row_data["couponUseEndAt"] = None
            row_data["couponUseTerm"] = 1  # TERM일 경우 사용 기간은 1로 설정
        else:
            # 기본값 처리 (DATE나 TERM이 아닌 경우)
            row_data["couponUseStartAt"] = None
            row_data["couponUseEndAt"] = None
            row_data["couponUseTerm"] = None

        #23. 사용 가능처 설정 (LIST필수)
        row_data["couponUseSiteStatus"] = (True if row_data.get("couponUseSiteStatus") in [True, "TRUE", 1, "1"] else False)

        #24. 사용 가능처 목록 설정 (가능처 설정 TRUE면 필수) - fixed_values로 관리

        #25. 사용 최소금액 설정 (LIST필수)
        # row_data["couponUseMinAmountStatus"] = "TRUE" if row_data.get("couponUseMinAmountStatus") in [True, "TRUE", 1, "1"] else "FALSE"
        row_data["couponUseMinAmountStatus"] = (True if row_data.get("couponUseMinAmountStatus") in [True, "TRUE", 1, "1"] else False)

        #26. 사용 최소 금액 처리 (최소금액 설정TRUE면 필수)
        row_data["couponUseMinAmount"] = "1000" if row_data["couponUseMinAmountStatus"] == True else None

        #27. 사용 최대 금액 설정(LIST필수)
        row_data["couponUseMaxAmountStatus"] = (True if row_data.get("couponUseMaxAmountStatus") in [True, "TRUE", 1, "1"] else False)

        #28. 사용 최대 금액 설정 여부(최대금액 설정TRUE면 필수)
        row_data["couponUseMaxAmount"] = "10000000" if row_data["couponUseMaxAmountStatus"] == True else None
        row_data["couponUseMinAmountStatus"] = (True if row_data.get("couponUseMinAmountStatus") in [True, "TRUE", 1, "1"] else False)
        #29. 파일 (쿠폰타입 BULK면 필수) - 업로드 파일을 올려야함.

        #쿠폰 안내 인덱스

        #30. 비밀번호 hash_value 처리 (필수)
        hash_value = hashlib.sha256(f"{couponType}{couponApplyMode}{couponApplyAmount}".encode()).hexdigest()
        row_data["hash"] = hash_value

        # Variable 시트에 데이터 추가
        full_row = [
            json.dumps(row_data.get(header, ""), ensure_ascii=False) 
            if isinstance(row_data.get(header, ""), (list, dict)) 
            else row_data.get(header, "")
            for header in variable_headers
        ]
        ws.append(full_row)


    # 파일 저장
    try:
        wb.save(excel_file_path)
        print(f"Combinations have been appended to the '{variable_sheet}' sheet in {excel_file_path}.")
    except Exception as e:
        print(f"Error while saving the file: {e}")

# 엑셀에서 데이터를 읽고 요청/응답을 기록하는 함수
def process_coupons_and_log_results(wb, file_path, variable_sheet, results_sheet):
    """
    기존 코드와 달리 DataFrame을 사용하지 않고 openpyxl로 직접 행(Row)을 순회하면서
    한 줄씩 읽어서 API 호출 후, 그 결과를 바로 results 시트에 기록하는 방식입니다.
    """
    # 엑셀 파일 로드 (openpyxl)
    # wb = load_workbook(file_path)

    # Variable 시트가 없으면 에러 발생
    if variable_sheet not in wb.sheetnames:
        raise ValueError(f"Sheet named '{variable_sheet}' not found in the file. Available sheets: {wb.sheetnames}")
    var_ws = wb[variable_sheet]

    # results 시트 준비
    if results_sheet not in wb.sheetnames:
        res_ws = wb.create_sheet(results_sheet)
        clear_sheet_content(res_ws)  # 기존 데이터 삭제 후 헤더 추가

        res_ws.append(["Requests", "Responses", "Status", "Coupon ID", "Match", "Match Details", "Mismatch Details", "Missing Details"])
    else:
        res_ws = wb[results_sheet]
        # 헤더가 없는 경우 추가
        if not any(res_ws[1]):
            res_ws .append(["Requests", "Responses", "Status", "Coupon ID", "Match", "Match Details", "Mismatch Details", "Missing Details"])
        else:
            # 헤더를 제외하고 모든 데이터 삭제
            print(f"Clearing data in '{results_sheet}' sheet except headers...")
            header = [cell.value for cell in res_ws[1]]  # 첫 번째 행만 저장
            res_ws.delete_rows(2, res_ws.max_row)  # 2행부터 마지막 행까지 삭제
            for col_num, value in enumerate(header, 1):
                res_ws.cell(row=1, column=col_num, value=value)

    url = "https://tcoupon.bodyfriend.com/api/coupon/v1/registration"
    headers = {
        "accept": "application/json;charset=UTF-8",
        "code": "COUPON",
        "key": "bodyfsdf#$asdfBdf3"
    }

    # Variable 시트의 헤더를 리스트로 추출 (1행)
    variable_headers = [cell.value for cell in var_ws[1] if cell.value]

    # 2행부터 마지막 행까지 처리
    last_row = var_ws.max_row

    # 콤비네이션에서 일정값만 가지고 오고 고정값이나 리스트값은 여기서 설정해서 바로 보내자!

    for row_idx in range(2, last_row + 1):
        # 한 행의 데이터를 dictionary 형태로 추출
        variable_data = {}
        for col_idx, header in enumerate(variable_headers, start=1):
            cell_value = var_ws.cell(row=row_idx, column=col_idx).value
            variable_data[header] = cell_value if cell_value != "" else None

        # 고정 값
        # 5번, 6번, 발급 전체 횟수 제한 처리 (선택)
        fixed_values = {
            "couponDupleIssueStatus": True,
            "couponDupleUseStatus": True,
            "couponIssueTotalLimit": "99999999",
        }

        # 리스트 값 추가
        list_values = {}


        # 8. 제품 카테고리 목록 couponGoodsCategoryList 처리
        if variable_data.get("couponGoodsCategoryStatus") == True:
            list_values["couponGoodsCategoryList"] = [
                {
                    "couponGoodsCategory": "MASSAGE",
                    "couponGoodsCategoryName": "바디프랜드"
                },
                {
                    "couponGoodsCategory": "LACLOUD",
                    "couponGoodsCategoryName": "라클라우드"
                },
                {
                    "couponGoodsCategory": "W-WATER",
                    "couponGoodsCategoryName": "W정수기"
                }
            ]

        
        # 10. 제품 목록  처리 couponGoodsList 처리
        if variable_data.get("couponGoodsStatus") == True:
            list_values["couponGoodsList"] = [
                {
                    "couponGoods": "1760",
                    "couponGoodsName": "LBF-750 - bfrms 제품 테스트"
                }
            ]


        # 12. 사용 가능 제외 제품 목록 couponGoodsNoneList 처리
        if variable_data.get("couponGoodsNoneStatus") == True:
            list_values["couponGoodsNoneList"] = [
                {
                    "couponGoods": "1760",
                    "couponGoodsName": "LBF-750 - bfrms 제품 테스트"
                }
            ]

        # 16. 발급가능처 목록 couponIssueSiteList 처리
        if variable_data.get("couponIssueSiteStatus") == True:
            list_values["couponIssueSiteList"] = [
                {
                    "couponSiteType": "ISSUE",
                    "couponSite": "MEMBERSHIP",
                    "couponSiteName": "MEMBERSHIP"
                }
            ]


        # 18. 고객 등급 목록 couponMemberGradeList 처리
        if variable_data.get("couponMemberGradeStatus") == True:
            list_values["couponMemberGradeList"] = [
                {
                    "couponMemberGrade": "FRIEND",
                    "couponMemberGradeName": "FRIEND"
                }
            ]

        # 24. 사용가능처 목록 couponUseSiteList 처리
        if variable_data.get("couponUseSiteStatus") == True:
            list_values["couponUseSiteList"] = [
                {
                    "couponSiteType": "ISSUE",
                    "couponSite": "MEMBERSHIP",
                    "couponSiteName": "MEMBERSHIP"
                }
            ]

        # 최종 request에 들어갈 payload
        payload_data = {**list_values, **variable_data, **fixed_values}


        # API 호출
        multipart_data = {
            "request": (None, json.dumps(
                {key: value for key, value in payload_data.items()},  # 문자열 변환 X
                ensure_ascii=False
            ), "application/json")
        }

        try:
            response = requests.post(url, headers=headers, files=multipart_data)
            time.sleep(1)

            request_str = json.dumps(payload_data, indent=2, ensure_ascii=False)
            try:
                response_json = response.json() or {}
                response_str = json.dumps(response_json, indent=2, ensure_ascii=False)
            except ValueError:
                print("Failed to parse JSON. Raw response text:")
                print(response.text)
                response_json = {}
                response_str = response.text

            status_code = response.status_code
            print(f"Response Status Code: {status_code}")
            print(f"[Row {row_idx - 1}/{last_row - 1}] Status Code: {status_code}")


            if status_code != 200:
                print(f"Unexpected status code. Response text: {response.text}")

            if not isinstance(response_json, dict):
                print("Warning: response_json is not a dictionary. Setting it to an empty dictionary.")
                response_json = {}

            if "data" not in response_json:
                print("'data' key is missing in response. Full response:")
                print(json.dumps(response_json, indent=2, ensure_ascii=False))

            coupon_id = response_json.get("data", {}).get("couponDetailInfo", {}).get("couponId", "N/A")

            # 요청값 vs 응답값 비교
            mismatches = []
            missing_keys = []
            matches = []

            response_data = response_json.get("data", {}).get("couponDetailInfo", {})
            
            # variable_headers에는 엑셀 헤더가 들어있음
            for key in variable_headers:
                request_value = str(payload_data[key]).strip().lower()
                response_value = str(response_data.get(key, "MISSING")).strip().lower()

                if response_value == "missing":
                    missing_keys.append(key)
                elif request_value != response_value:
                    mismatches.append(f"{key}: {request_value} != {response_value}")
                else:
                    matches.append(key)

            if not mismatches and not missing_keys:
                result = "perfect"
            elif not mismatches and missing_keys:
                result = "matched"
            else:
                result = "false"

            mismatch_details = ", ".join(mismatches) if mismatches else "None"
            missing_details = ", ".join(missing_keys) if missing_keys else "None"
            match_details = ", ".join(matches) if matches else "None"

        except requests.exceptions.RequestException as e:
            # 요청 실패 처리
            print(f"Request failed: {e}")
            result = "FAIL"
            response_str = str(e)
            coupon_id = "N/A"
            mismatch_details = "Request Exception"
            missing_details = "None"
            match_details = response.text if hasattr(response, 'text') else "No response text"

        except AttributeError as e:
            # response_json에서 발생한 AttributeError 처리
            print(f"AttributeError occurred: {e}")
            result = "FAIL"
            response_str = f"Error: {str(e)}"
            coupon_id = "N/A"
            mismatch_details = "Invalid response structure"
            missing_details = "None"
            match_details = response.text if hasattr(response, 'text') else "No response text"

        res_ws.append([
            request_str, 
            response_str, 
            status_code, 
            coupon_id, 
            result, 
            match_details, 
            mismatch_details, 
            missing_details])

        time.sleep(1)
        
        # 실시간 저장 및 저장여부 노출
        wb.save(file_path)
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"Row {row_idx - 1} processed. CouponID: {coupon_id}, Match: {result}")

        # n번째(예: 1000행)마다 스냅샷 저장
        if (row_idx - 1) % 500 == 0:
            wb.save("coupontest_variable_snapshot.xlsx")
            print("Saved a snapshot for checking progress.")


try:
    wb = load_workbook(excel_file_path)

    # 함수 실행
    generate_combinations(wb, variable_list_sheet, variable_sheet)
    process_coupons_and_log_results(wb, excel_file_path, variable_sheet, results_sheet)
except KeyboardInterrupt:
    print("KeyboardInterrupt detected! Cleaning up and safely exiting...")
finally:
    try:
        # 엑셀 파일 안전 저장
        wb.save(excel_file_path)
        wb.close()
        print(f"Excel file '{excel_file_path}' has been safely saved.")
    except Exception as e:
        print(f"Error during cleanup: {e}")